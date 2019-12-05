import cognitive_face as CF
from global_variables import personGroupId, Key, endpoint
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import cell
import time


CF.BaseUrl.set(endpoint)
CF.Key.set(Key)

#get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = "reports.xlsx")
sheet = wb.get_sheet_by_name('Cse15')

def getDateColumn():
	for i in range(1, 100):
		col =i
		#print(sheet.cell(row=1 ,column=col).value)
		if (sheet.cell(row=1,column=col).value == currentDate):
			return col
			


conn = sqlite3.connect("Face-DataBase")
c = conn.cursor()

attend = [0 for i in range(60)]	

currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, 'Cropped_faces')
for filename in os.listdir(directory):
	if filename.endswith(".jpg"):
		imgurl = urllib.request.pathname2url(os.path.join(directory, filename))
		res = CF.face.detect(imgurl)
		if len(res) != 1:
			print("No face detected.")
			continue
			
		faceIds = []
		for face in res:
			faceIds.append(face['faceId'])
		res = CF.face.identify(faceIds, personGroupId)
		print(filename)
		print(res)
		for face  in res:
			if not face['candidates']:
				print("Unknown")
			else:
				personId = face['candidates'][0]['personId']
				c.execute("SELECT * FROM Students")
				record = c.fetchall()
				for r in record:
					if r[3]==personId:
						attend[r[0]]+=1
						print(r[1]+" recognised")
		time.sleep(1)
		
for row in range(3,60):
	rn = str(sheet.cell(row,column=1).value)
	if rn is not None:
		if(rn == "None"):
			break
		rn = rn[-2:]
		if attend[int(rn)] != 0:
			col = getDateColumn()
			sheet.cell(row, column=col).value = 1

wb.save(filename = "reports.xlsx")	 	
#currentDir = os.path.dirname(os.path.abspath(__file__))
#imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
#res = CF.face.detect(imgurl)
#faceIds = []
#for face in res:
 #   faceIds.append(face['faceId'])

#res = CF.face.identify(faceIds,personGroupId)
# for face in res:
#     personName = CF.person.get(personGroupId, face['candidates']['personId'])
#     print personName
#print res
