from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import Cell
import time

attend=range(1,60)

#get current date
currentDate = time.strftime("%d_%m_%y")

def getDateColumn2():
	for i in range(1, 100):
		#col = get_column_letter(i)
		#print(sheet.cell(row=1 ,column=i).value)
		if (sheet2.cell(row=1 ,column=i).value) == currentDate:
			return i

def getDateColumn1():
	for i in range(1, 100):
		#col = get_column_letter(i)
		#print(sheet.cell(row=1 ,column=i).value)
		if (sheet1.cell(row=1 ,column=i).value) == currentDate:
			return i

#adding the final attendance from the report

wb1 = load_workbook(filename = "result.xlsx")
sheet1 = wb1.get_sheet_by_name('Cse15')


wb2 = load_workbook(filename = "reports.xlsx")
sheet2 = wb2.get_sheet_by_name('Cse15')


for row in range(3,60):
	rn = str(sheet2.cell(row,column=1).value)
	if rn is not None:
		if(rn == "None"):
			break
		rn = rn[-2:]
		if attend[int(rn)] != 0:
			col = getDateColumn2()
			col2=col+1
			if((sheet2.cell(row, column=col).value == 1)&(sheet2.cell(row, column=col2).value == 1)):
				c = getDateColumn1()
				sheet1.cell(row, column=c).value = 1


wb2.save(filename = "reports.xlsx")
wb1.save(filename = "result.xlsx")
