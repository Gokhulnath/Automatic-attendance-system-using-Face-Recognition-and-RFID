from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import Cell
import time
import os
import sqlite3

#database connection
conn = sqlite3.connect('Face-DataBase')
c = conn.cursor()

#get current date
currentDate = time.strftime("%d_%m_%y")

#create a workbook and add a worksheet
if(os.path.exists('./reports.xlsx')):
    wb = load_workbook(filename = "reports.xlsx")
    sheet = wb.get_sheet_by_name('Cse15')
    # sheet[ord() + '1']
    for col_index in range(3,100,2):
    	col = col_index
    	if sheet.cell(row=1,column=col).value is None:
    		col2 = col_index - 2
    		if sheet.cell(row=1,column=col2).value != currentDate:
    			sheet.cell(row=1,column=col_index).value = currentDate
    		break

    #saving the file
    wb.save(filename = "reports.xlsx")
    	
else:
    wb = Workbook()
    dest_filename = 'reports.xlsx'
    c.execute("SELECT * FROM Students ORDER BY Roll ASC")
    
    #creating worksheet and giving names to column
    ws1 = wb.active
    ws1.title = "Cse15"
    ws1.append(('Roll Number', 'Name', currentDate))
    ws1.append(('', '', ''))

    #entering students information from database
    while True:
        a = c.fetchone()
        if a == None:
            break
        else:
            ws1.append((a[2], a[1]))

    #saving the file
    wb.save(filename = dest_filename)
    
    
