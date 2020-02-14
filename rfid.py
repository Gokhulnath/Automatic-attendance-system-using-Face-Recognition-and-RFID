from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import cell
import time
import re
import serial
from serial import Serial
import signal
from contextlib import contextmanager

currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = "reports.xlsx")
sheet = wb.get_sheet_by_name('Cse15')

attend = [0 for i in range(60)]	


class TimeoutException(Exception): pass

@contextmanager
def time_limit(seconds):
    def signal_handler(signum, frame):
        raise TimeoutException("Timed out!")
    signal.signal(signal.SIGALRM, signal_handler)
    signal.alarm(seconds)
    try:
        yield
    finally:
        signal.alarm(0)

def rfidread():
	while(1):
		ser_bytes = ser.read(size=8)
		#print(ser_bytes)
		ser_bytes=str(ser_bytes)
		ser_bytes=ser_bytes[24:26]
		ser_bytes=re.sub('[^0-9]', '', ser_bytes)
		ser_bytes=int('0'+ser_bytes)
		print(ser_bytes)
		if(ser_bytes!=0):
			attend[ser_bytes]+=1


def getDateColumn():
	for i in range(1, 100):
		#col = get_column_letter(i)
		#print(sheet.cell(row=1 ,column=i).value)
		if (sheet.cell(row=1 ,column=i).value) == currentDate:
			return i



ser = serial.Serial(port='/dev/ttyUSB0', baudrate=57600, bytesize=serial.EIGHTBITS, parity=serial.PARITY_NONE, timeout=2)
try:
    ser.isOpen()
    print("serial port is open")
except:
    print("error")
    exit()

if(ser.isOpen()):
	try:
		with time_limit(10):
			rfidread()
	except TimeoutException as e:
		print("Timed out!")
else:
    print("cannot open serial port")
ser.close()

for row in range(3,60):
	rn = str(sheet.cell(row,column=1).value)
	if rn is not None:
		if(rn == "None"):
			break
		rn = rn[-2:]
		if attend[int(rn)] != 0:
			col = getDateColumn()+1
			sheet.cell(row, column=col).value = 1


wb.save(filename = "reports.xlsx")
